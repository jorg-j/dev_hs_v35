"""
Manage the mainline for report generation
"""

import datetime
import json
import logging
import platform
import sys
from inspect import getsourcefile  # block
from os import path  # block

# pylint: disable=unused-import
import openpyxl
import pytz
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo


#### SOF


def report_header(workbook, hs_id):
    """
    Setup the report header

    -------------------------------------------------
    |             report_heading                |
    |Hyperscience ID      | 789                     |
    |Report Date          | d/m/y I:M p             |
    -------------------------------------------------
    """
    worksheet = manage_worksheet(workbook, ReportDefaults.sheet_main)

    # starting positions
    pos = PositionStore(row=3, col=2)

    # Create the report headline
    cell = worksheet.cell(column=pos.col, row=pos.row)
    cell.value = ReportDefaults.title_main
    cell.font = ReportFont.h1
    cell.border = ReportBoarder.thin_border

    # Merge the heading cell
    merge_wrap(worksheet, "B3:E3", wrap=False)

    pos.row = 2


    # Create Submission Summary Heading
    pos = section_header(worksheet, ReportDefaults.report_heading, pos)

    # Setup each detail

    def write_sub_sum_items(worksheet, title, value, pos):
        """
        Write out the sub summary items as they have special formatting
        which has 2 cells cover 4 columns
        """
        cell = worksheet.cell(column=pos.col, row=pos.row)
        cell.value = title
        cell.font = ReportFont.body_bold
        cell.border = ReportBoarder.thin_border
        cell.alignment = Alignment(horizontal="left")

        cell = worksheet.cell(column=pos.col + 2, row=pos.row)
        cell.value = value
        cell.border = ReportBoarder.thin_border
        cell.alignment = Alignment(horizontal="left")
        cell.fill = ReportFill.green

        merge_wrap(worksheet, f"B{pos.row}:C{pos.row}", align_h="left")
        merge_wrap(worksheet, f"D{pos.row}:E{pos.row}", align_h="left")
        return worksheet


    ## Hyperscience ID
    write_sub_sum_items(worksheet, "HYPERSCIENCE ID", hs_id, pos)
    pos.row = 1

    ## REPORT DATE
    timezone = pytz.timezone("Australia/Sydney")
    right_now = datetime.datetime.now(timezone)
    formatted_time = right_now.strftime("%d/%m/%y %I:%M %p")
    timestamp = right_now.strftime("%y%m%d_%H%M")
    write_sub_sum_items(worksheet, "REPORT DATE", formatted_time+' EST', pos)

    pos.row = 1

    blank_row(worksheet, pos)

    auto_width(worksheet, max_width=200)
    return workbook, pos, timestamp


def section_header(worksheet, heading, pos):
    """
    Write out a section heading
    -------------------------------------------------
    |               heading goes here               |
    -------------------------------------------------
    """
    # Create Submission Summary Heading
    cell = worksheet.cell(column=pos.col, row=pos.row)
    cell.value = heading
    cell.font = ReportFont.h2_white
    cell.border = ReportBoarder.thin_border
    cell.fill = ReportFill.blue
    cell.alignment = Alignment(horizontal="center")

    # Merge the heading cell
    merge_wrap(worksheet, f"B{pos.row}:E{pos.row}", wrap=False)
    pos.row = 1
    return pos



def run_reporting(data):
    """
    Reporting Main line


    Note: the pos variable will maintain the current position in the reports main page
    """
    workbook = init_workbook()

    # Write the disclaimer to excel
    disclaimer(workbook)


    workbook, pos, timestamp = report_header(
        workbook, hs_id=data.get("hs_submission_id", 0)
    )

    pos.row = 1

    # Table Func call here # TODO


    worksheet = manage_worksheet(workbook, ReportDefaults.sheet_main)
    auto_width(worksheet)

    SaveWorkbook(workbook, hs_id=data.get("hs_submission_id", 0), timestamp=timestamp).disk()


#### EOF