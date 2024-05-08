"""
Utilities to support reporting

"""


import copy
import logging

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(
    encoding="utf-8",
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s: %(name)s: Line: %(lineno)s - %(funcName)s(): Splunk Send Start > %(message)s",  # pylint: disable=line-too-long
)


# pylint: disable=redefined-outer-name
# pylint: disable=broad-exception-raised
# pylint: disable=broad-exception-caught
# pylint: disable=missing-function-docstring
def log_info(text):
    logging.info(text)


def log_warn(text):
    logging.warning(text)


def log_debug(text):
    logging.debug(text)


# pylint: enable=missing-function-docstring

#### SOF


def merge_wrap(worksheet, range_value, wrap=True, align_v="center", align_h="center"):
    """
    merge cell ranges
    """
    worksheet.merge_cells(range_value)
    cell = worksheet[range_value.split(':')[0]]
    cell.alignment = Alignment(vertical=align_v, horizontal=align_h)
    alignment = copy.copy(cell.alignment)
    alignment.wrapText = wrap
    cell.alignment = alignment


def merge_rows(worksheet, start_row, end_row, column_letter="B"):
    """
    Merge Cells together which contain the same values
    """

    # Collect all values from the column range
    column_values = []

    # collect values
    for item in range(start_row, end_row):
        cell_value = worksheet.cell(row=item, column=column_letter)
        column_values.append(cell_value)

    result = []

    # Exit early if no data
    if not column_values:
        return

    # set the first item in the column_values list so it knows what to start comparing to
    current_item = column_values[0]
    current_pos = 0
    for i, item in enumerate(column_values):
        # if row values are different then register the prior range
        if item != current_item:
            end_pos = i - 1 if i - 1 > 0 else 0
            result.append((current_pos, end_pos))
            current_pos = i
            current_item = item

    # allows the end of the range to be merged if needed
    if current_item is not None:
        if current_pos != i:
            result.append((current_pos, len(column_values) - 1))

    for item in result:
        # Ensure the positions arent the same
        if item[0] != item[1]:
            # perform the merge
            start = item[0] + start_row
            end = item[1] + start_row
            merge_wrap(worksheet, f"{column_letter}{start}:{column_letter}{end}")


def auto_width(worksheet, max_width=41):
    """
    Adjusts the column widths of a given worksheet based on the length of the cell values.

    :param worksheet: auto_width takes this worksheet object as input and 
    adjusts the column widths of the
    worksheet based on the length of the data in each cell. The function then returns the modified
    worksheet object
    :return: the worksheet object with adjusted column widths.
    """
    log_info("Adjusting column widths")
    for column_cells in worksheet.columns:
        max_length = 0
        max_font = 10
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_value = str(cell.value)
                cell_font = cell.font.sz
            except:
                cell_value = ""
                cell_font = 10
            cell_length = len(cell_value)
            if cell_length > max_length:
                max_length = cell_length
            try:
                if cell_font > max_font:
                    max_font = cell_font
            except:
                pass
        if max_length != 0:
            adjusted_width = (max_length) * 1.2
            adjusted_width *= max_font / 10
            adjusted_width = min(adjusted_width, max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return worksheet


#### EOF
