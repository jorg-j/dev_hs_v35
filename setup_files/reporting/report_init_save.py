"""
Manage Saving of workbook to either
Disk
Blob Storage
OBP API Upload
"""

import base64
import io
import logging
import copy
import openpyxl
from dataclasses import dataclass

logging.basicConfig(
    encoding="utf-8",
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s: %(name)s: Line: %(lineno)s - %(funcName)s(): Splunk Send Start > %(message)s",  # pylint: disable=line-too-long
)


# pylint: disable=redefined-outer-name
# pylint: disable=broad-exception-raised
# pylint: disable=broad-exception-caught
# pylint: disable=missing-function-docstring
def log_info(text):
    logging.info(text)


def log_warn(text):
    logging.warning(text)


def log_debug(text):
    logging.debug(text)


# pylint: enable=missing-function-docstring


#### SOF

@dataclass
class ReportDefaults:
    sheet_main = "HyperScience_Report"
    title_main = "Title Goes Here"
    report_heading = "Report heading goes here"




def init_workbook():
    """
    Init the workbook and its required pages
    """
    workbook = openpyxl.Workbook()

    workbook_default_sheets = [ReportDefaults.sheet_main]

    for s in workbook_default_sheets:
        worksheet = manage_worksheet(workbook, s)

    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    return workbook    



def manage_worksheet(workbook, sheet_name):
    """
    Manages a worksheet in a given workbook by either retrieving an existing worksheet or
    creating a new one with a specified name.

    :param workbook: The parameter 'workbook' is a variable that represents a workbook object in openpyxl.
    :param sheet_name: The name of the worksheet that needs to be managed or created
    :return: a worksheet object, either an existing worksheet with the given name or a newly created
    worksheet with the given name.
    """
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

    else:
        worksheet = workbook.create_sheet(sheet_name)
    workbook.active = workbook.index(worksheet)
    return worksheet



class SaveWorkbook:
    """
    Saves the workbook to the appropriate location

    Usage:
        SaveWorkbook(workbook, workbook, timestamp).disk()
        SaveWorkbook(workbook, workbook, timestamp).blob()
        SaveWorkbook(workbook, workbook, timestamp).api(doc_guid, domain, api_key)
        SaveWorkbook(workbook, workbook, timestamp).both(doc_guid, domain, api_key)

    """
    def __init__(self, workbook, hs_id, timestamp):
        self.workbook = workbook
        self.file_name = f"Report_{hs_id}_{timestamp}.xlsx"
        self.submission_id = hs_id


        self.buffer_encoding = None
        self.buffer_raw = None
        self.encoded_workbook = None

    def _store(self):
        """
        Takes the workbook puts it in the io buffer
        """
        log_info("Writing Workbook to buffer")
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        self.buffer_encoding = copy.deepcopy(buffer)
        self.buffer_raw = copy.deepcopy(buffer)
        log_info("Write Complete")


    def _wb_encode(self):
        log_info("Encoding Workbook")
        self.encoded_workbook = base64.b64encode(self.buffer_encoding.read()).decode()
        log_info("Workbook Encoded")


    def disk(self):
        """
        Save workbook to disk
        """
        self.workbook.save(self.file_name)

    def both(self, doc_guid, domain, api_key):
        """
        Upload to both API and Blob storage
        """
        self.blob()
        self.api(doc_guid, domain, api_key)

    def api(self, doc_guid, domain, api_key):
        """
        Saves the workbook to the OBP API
        """

        if self.buffer_encoding is None:
            self._store()
        if self.encoded_workbook is None:
            self._wb_encode()

        log_info(
            f"If there is a domain in this message the key fetch has been successful: {domain}"
        )

        ## TODO implement API upload

        response_code = 501

        if str(response_code).startswith("2"):
            log_info(
                f"Submission: {self.submission_id} was successful with response: {response_code}"
            )
        else:
            log_warn(
                f"Submission: {self.submission_id} has failed with response: {response_code}"
            )
            raise Exception(f"API Upload failed with Response Code {response_code}")


    def blob(self):
        """
        Saves the workbook to blob storage
        """
        log_info("Blob storage requested")

        try:
            from flows_sdk.types import StoreBlobRequest # pylint: disable=import-outside-toplevel

            log_info("HS block instance StoreBlobRequest init")
        except ImportError as exc:
            log_warn("Unable to import StoreBlobRequest")
            raise Exception(f"Import Error - {exc}") # pylint: disable=raise-missing-from

        # Write the report to io buffer
        if self.buffer_raw is None:
            self._store()

        # Write the io buffer to blob storage
        log_info("Writing report to blob storage: HS Block")
        # pylint: disable=undefined-variable
        blob_file = _hs_block_instance.store_blob(
            StoreBlobRequest(
                name=self.file_name,
                content=self.buffer_raw.read(),
            )
        )
        # pylint: enable=undefined-variable

        log_info("Writing report to blob storage: Complete")

        # Write the blob file uuid to logs
        abs_path = f"/api/block_storage/{blob_file.uuid}/download"
        log_info(f"Report saved to blob storage: {abs_path}")

#### EOF
