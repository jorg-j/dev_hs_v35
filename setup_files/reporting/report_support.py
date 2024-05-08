"""
Support functions for creating report line and items
"""

from urllib.parse import quote
import logging

import sys

from dataclasses import dataclass
from inspect import getsourcefile  # block
from os import path  # block
import copy

# pylint: disable=unused-import
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo

current_dir = path.dirname(path.abspath(getsourcefile(lambda: 0)))
sys.path.insert(0, current_dir[: current_dir.rfind(path.sep)])

from reporting.formatting import *  # pylint: disable=wrong-import-position,import-error # block

logging.basicConfig(
    encoding="utf-8",
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s: %(name)s: Line: %(lineno)s - %(funcName)s(): Splunk Send Start > %(message)s",  # pylint: disable=line-too-long
)


# pylint: disable=redefined-outer-name
# pylint: disable=broad-exception-raised
# pylint: disable=broad-exception-caught
# pylint: disable=missing-function-docstring
def log_info(text):
    logging.info(text)


def log_warn(text):
    logging.warning(text)


def log_debug(text):
    logging.debug(text)


# pylint: enable=missing-function-docstring


#### SOF

def generate_mail_link(cell, cell_text, to, subject, body):
    """
    Generates a mail link using the mailto protocol
    Send in 
    cell - where this link will be
    cell_text - as the "click me"
    to - recipient
    subject - mail subject
    body - mail body
    """

    cell.value = cell_text
    subject_clean = quote(subject)
    body_clean = quote(body)
    mail_to = f"mailto:{to}?subject={subject_clean}&body={body_clean}"
    cell.hyperlink = mail_to
    cell.font = ReportFont.body_bold
    cell.fill = ReportFill.orange

def write_standard_body(cell, value):
    '''
    writes standard text to report as standard body
    '''
    cell.value = value
    cell.font = ReportFont.body
    cell.border = ReportBoarder.thin_border
    cell.alignment = Alignment(horizontal="left")

def write_row_values(worksheet, item_list, row, **kwargs):
    """
    Writes a list of items to a worksheet in a specified row and column.
    Note this writes a horizontal list

    worksheet: the worksheet object
    item_list: list of items to be written to the worksheet
    row: starting row where items will be written
    column: optional - column to write the values into, default = 2
    heading: optional - is this row a heading, default = False
    table: optional - is the row part of a table, default = True

    border: optional - boolean for cell borders, default = True
    wordwrap: optional - boolean for cell value wordwrapping, default = True
    """
    column = kwargs.get("column", 2)
    wordwrap = kwargs.get("wordwrap", True)

    is_heading = kwargs.get("heading", False)
    is_table = kwargs.get("table", True)

    # Set default font
    if is_heading:
        font = kwargs.get("font", ReportFont.h3)
    else:
        font = kwargs.get("font", ReportFont.body)

    log_debug(f"Writing row data: {str(item_list)}")
    for i, item in enumerate(item_list, start=column):
        cell = worksheet.cell(row=row, column=i)
        cell.value = item

        # Format table by adding border and alignment left
        if is_table:
            cell.border = ReportBoarder.thin_border
            cell.alignment = Alignment(horizontal="left")
        
        # Apply wordwrap to the cell, prevents truncation
        if wordwrap:
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment

        # Apply either a heading format or standard font
        if is_heading:
            cell.fill = ReportFill.lightblue
            cell.font = font
        else:
            cell.font = font

def disclaimer(workbook):
    """
    Puts the disclaimer in the report
    """

    worksheet = manage_worksheet(workbook, ReportDefaults.sheet_main)

    disclaimer = "Disclaimer: This document contains confidential information. Any unauthorised use, disclosure, copying or distribution of this document is strictly prohibited. If you have received this document in error, please notify the sender immediately and delete it from your system."  # pylint: disable=line-too-long

    cell = worksheet.cell(row=1, column=2)
    cell.value = disclaimer
    cell.font = ReportFont.body_italics
    cell.alignment = Alignment(horizontal="center")
    alignment = copy.copy(cell.alignment)
    alignment.wrapText = True
    merge_wrap(worksheet, "B1:D1")
    # Adjust the row height so it doesnt look funky with all that text
    worksheet.row_dimensions[3].height = 43

class PositionStore:
    def __init__(self, row=2, col=2):
        self._row = row
        self._col = col
        self.offset = 0
        self.offset_col = 0

    @property
    def row(self):
        return self._row
    
    @row.setter
    def row(self, amount=1):
        self.offset += amount
        self._row += amount

    @property
    def col(self):
        return self._col
    
    @col.setter
    def col(self, amount=1):
        self.offset_col += amount
        self._col += amount




#### EOF