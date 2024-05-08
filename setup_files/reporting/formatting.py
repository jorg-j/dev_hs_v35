"""
Managing the formatting functions for the report
"""
import logging
import sys
from dataclasses import dataclass
from inspect import getsourcefile  # block
from os import path  # block

# pylint: disable=unused-import
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo

current_dir = path.dirname(path.abspath(getsourcefile(lambda: 0)))
sys.path.insert(0, current_dir[: current_dir.rfind(path.sep)])

from reporting.excel_utilities import *  # pylint: disable=wrong-import-position,import-error # block

logging.basicConfig(
    encoding="utf-8",
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s: %(name)s: Line: %(lineno)s - %(funcName)s(): Splunk Send Start > %(message)s",  # pylint: disable=line-too-long
)


# pylint: disable=redefined-outer-name
# pylint: disable=broad-exception-raised
# pylint: disable=broad-exception-caught
# pylint: disable=missing-function-docstring
def log_info(text):
    logging.info(text)


def log_warn(text):
    logging.warning(text)


def log_debug(text):
    logging.debug(text)


# pylint: enable=missing-function-docstring


#### SOF


@dataclass
class Palette:
    """
    Color Palette for the report
    Note: the blues have been checked for colorblindness
        they all show differently under all variants of colorblindness
    """

    black = "00000000"
    red = "FF0000"
    orange = "FFBF00"
    green = "C6EFCE"
    white = "FFFFFF"
    grey = "BEBEBE"
    gray = grey
    none = ""

    # Blues
    blue = "0000FF"
    frenchblue = "0072bb"
    reportblue = "B4C6E7"
    hawkesblue = "C5D9F1"
    hyperlink = "0563C1"


@dataclass
class ReportBoarder:
    """
    Friendly names for cell boarders
    """

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    top_bottom_left_border = Border(
        left=Side(style="thin"),
        right=None,
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    top_bottom_right_border = Border(
        left=None,
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    top_bottom_border = Border(
        left=None,
        right=None,
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    top_border = Border(
        left=None,
        right=None,
        top=Side(style="thin"),
        bottom=None,
    )

    bottom_border = Border(
        left=None,
        right=None,
        top=None,
        bottom=Side(style="thin"),
    )

    left_border = Border(
        left=Side(style="thin"),
        right=None,
        top=None,
        bottom=None,
    )

    right_border = Border(
        left=None,
        right=Side(style="thin"),
        top=None,
        bottom=None,
    )

    top_left_border = Border(
        left=Side(style="thin"),
        right=None,
        top=Side(style="thin"),
        bottom=None,
    )

    top_right_border = Border(
        left=None,
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=None,
    )

    bottom_left_border = Border(
        left=Side(style="thin"),
        right=None,
        top=None,
        bottom=Side(style="thin"),
    )

    bottom_right_border = Border(
        left=None,
        right=Side(style="thin"),
        top=None,
        bottom=Side(style="thin"),
    )


@dataclass
class ReportFont:
    """
    Standard Fonts used within the report

    cell.font = ReportFont.black
    """

    default_font = "Calibri"
    h1_size = 20
    h2_size = 11
    h3_size = 9
    body_size = 9
    alert_size = 9

    hyperlink = Font(color=Palette.hyperlink, underline="single")
    red = Font(color=Palette.red)
    black = Font(color=Palette.black)

    h1 = Font(name=default_font, size=h1_size, color=Palette.black, bold=True)

    h2 = Font(name=default_font, size=h2_size, color=Palette.black, bold=True)

    h2_white = Font(name=default_font, size=h2_size, color=Palette.white, bold=True)

    h3 = Font(name=default_font, size=h3_size, color=Palette.black, bold=True)

    body = Font(name=default_font, size=body_size, color=Palette.black, bold=False)

    body_bold = Font(name=default_font, size=body_size, color=Palette.black, bold=True)

    body_italics = Font(
        name=default_font, size=body_size, color=Palette.black, italic=True
    )

    alert = Font(name=default_font, size=alert_size, color=Palette.red, bold=True)


@dataclass
class ReportFill:
    """
    Pattern fills used across the report

    cell.fill = ReportFill.blue
    """

    orange = PatternFill(
        start_color=Palette.orange, end_color=Palette.orange, fill_type="solid"
    )

    green = PatternFill(
        start_color=Palette.green, end_color=Palette.green, fill_type="solid"
    )
    red = PatternFill(start_color=Palette.red, end_color=Palette.red, fill_type="solid")

    white = PatternFill(
        start_color=Palette.white, end_color=Palette.white, fill_type="solid"
    )

    # used as the section heading blue
    blue = PatternFill(
        start_color=Palette.reportblue,
        end_color=Palette.reportblue,
        fill_type="solid",
    )

    # used as the heading blue tables
    lightblue = PatternFill(
        start_color=Palette.hawkesblue,
        end_color=Palette.hawkesblue,
        fill_type="solid",
    )


def color_cell(cell, color):
    """
    Set the cell color
    """
    try:
        color_code = getattr(Palette, color.lower())
        if color_code:
            fill = PatternFill(
                start_color=color_code, end_color=color_code, fill_type="solid"
            )
            cell.fill = fill
    except:
        raise Exception(
            f"Color: {color} is not found in Palette"
        )  # pylint: disable=raise-missing-from


def blank_row(worksheet, position):
    """
    Sets the row to be blank
    B-E cols will be merges and cell fill white.
    So it looks nice
    """
    cell = worksheet.cell(column=position.col, row=position.row)
    color_cell(cell, "white")
    merge_wrap(worksheet, f"B{position.row}:E{position.row}")


def hyperlink_internal(cell, sheet, column, row):
    """
    Creates a hyperlink in a cell which links to a specific cell in a specific sheet

    cell: The cell object we want to turn into a hyperlink
    sheet: The name of the sheet where the hyperlink directs to
    column: The column letter ("A", "B"...) where the hyperlink directs to
    row: The row number the hyperlink directs to
    """
    text = cell.value
    cell_value = f'=HYPERLINK("#{sheet}!{column}{row}", "{text}")'
    cell.value = cell_value
    cell.style = "Hyperlink"
    cell.border = ReportBoarder.thin_border


def hyperlink_external(cell, url, border=True):
    """
    Sets a hyperlink and font style for a given cell in a spreadsheet.

    :param cell: The cell parameter is the cell object in which the
    hyperlink will be inserted.
    :param url: The URL the hyperlink should point to.
    """
    cell.hyperlink = url
    cell.font = ReportFont.hyperlink

    cell.style = "Hyperlink"
    if border:
        cell.border = ReportBoarder.thin_border


#### EOF
